from openpyxl import load_workbook


def xlsx_info(input_file: str):
    wb = load_workbook(input_file)
    ws = wb.active


    row_count = 0
    for row in range(1, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        if cell_value is not None and str(cell_value).strip():
            row_count += 1


    start_row = 0
    urls_col = 0
    for col in range(1, row_count + 1):
        for row in range(1, row_count + 1):
            cell_url = ws.cell(row=row, column=col).value
            cell_value = ws.cell(row=row, column=col+1).value
            if isinstance(cell_url, str) and cell_url.startswith('https://') and not cell_value:
                start_row = row
                urls_col = col
                break

    if start_row == 0:
        return [[start_row, urls_col], 0]

    return [[start_row, urls_col], row_count-start_row + 1]


def save_cell(input_file: str, cell: list[int, int], value: str):
    wb = load_workbook(input_file)
    ws = wb.active
    ws.cell(row=cell[0], column=cell[1] + 1).value = value
    wb.save(input_file)


def cell_value(input_file: str, cell: list[int, int]):
    wb = load_workbook(input_file)
    ws = wb.active

    return ws.cell(row=cell[0], column=cell[1]).value


